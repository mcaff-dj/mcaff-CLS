"""Shared per-brand report state + low-level helpers.

PowerShell dot-sources Generate-Report.ps1's variables/functions into gen-insights.ps1,
gen-weekly.ps1, gen-monthly.ps1 and gen-panels.ps1 (all share one script scope). Python has
no equivalent, so this module holds the small set of pure helpers those files all use
(HTML/JSON escaping, month/year formatting) plus a plain `Ctx` object that generate_report.py
builds once per brand and every other module receives explicitly - the direct replacement
for that shared script scope.
"""
import math
import re
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent


def ci_key(value, cache):
    """Resolves `value` to the first-seen casing recorded in `cache` (a plain dict mapping
    casefolded string -> first-seen original string). PowerShell's default @{}/[ordered]@{}
    hashtable compares string keys case-insensitively but preserves the casing of whichever
    variant was inserted first - two rows spelling the same category "Product not Sealed"
    and "product NOT sealed" collapse into one bucket there. Python dicts are always
    case-sensitive, so every grouping-by-text-value operation ported from a PowerShell
    hashtable must pre-resolve its key through this function first to match exactly.
    Pass a fresh `cache = {}` per independent grouping operation (matching one `@{}`
    instantiation in the original) - a single cache shared across unrelated groupings would
    let an unrelated part of the report "lock in" a casing before this one sees it, which
    isn't how independently-scoped PowerShell hashtables behave."""
    if not isinstance(value, str):
        return value
    cf = value.casefold()
    if cf in cache:
        return cache[cf]
    cache[cf] = value
    return value

# The older KYC raw-dump sheet (see brands.py "secondary") worded some Query Category
# values differently than the primary sheet - normalized here so both sides of a merged
# report count as one category instead of splitting the same complaint type into two rows.
CAT_NORM_MAP = {
    "Reattempt Request/ Fake update": "Fake update",
    "Pincode non Serviceable": "Pincode not serviceable",
    "Lost order/Destroyed/Damaged": "Lost/Damaged/Destroyed",
    "Marked Delivered but customer did not received the order": "Marked Delivered but customer did not receive order",
    "Dull-skin issue": "Dull skin issue",
}
# Same reasoning as CAT_NORM_MAP above, but for Delivery Partner Name: the older KYC
# raw-dump sheet has messy operational sub-labels (surface/air legs, direct/hyphen
# routing codes, brand-specific suffixes) for couriers the primary sheet already
# reports under clean canonical names.
PARTNER_NORM_MAP = {
    "Blue Dart Air": "Blue Dart",
    "Blue Dart Surface": "Blue Dart",
    "Bluedart": "Blue Dart",
    "Bluedart brands 500 g Surface": "Blue Dart",
    "Bluedart Surface - Select  500gm": "Blue Dart",
    "Bluedart Surface - Select 500gm": "Blue Dart",
    "Bluedart Surface 500 gms- Select": "Blue Dart",
    "Cuberooteeine": "PurpleDrone",
    "Purpledrone_mCaff": "PurpleDrone",
    "Delhivery Air": "DELHIVERY",
    "DELHIVERY_SMYTTEN": "DELHIVERY",
    "DLSRF_Direct": "DELHIVERY",
    "Dlv_Direct_Air": "DELHIVERY",
    "HYP_DELHIVERY": "DELHIVERY",
    "SR_Delhivery": "DELHIVERY",
    "DTDC Surface": "DTDC",
    "DTDC_Surface_Direct": "DTDC",
    "Ekart Logistics Surface": "Ekart",
    "Elasticrun_direct_M": "ElasticRun",
    "Pidge_Omnivio": "Pidge",
    "Pikndel_M_SDD": "Pikndel",
    "Shadowfax Surface": "Shadowfax",
    "SHADOWFAX_ESSENTIAL": "Shadowfax",
    "Shadowfax_M_NDD": "Shadowfax",
    "Shadowfax_M_SDD": "Shadowfax",
    "Fedex Air": "Fedex",
    "SR_Fedex_Courier": "Fedex",
    "XBSRF_ Air_Direct": "Xpressbees",
    "XBSRF_Direct": "Xpressbees",
    "XBSRF_Direct_NDD": "Xpressbees",
    "xpressbees": "Xpressbees",
    "Xpressbees Surface": "Xpressbees",
    "na": "(blank)",
}

# The sheet's "Order Month" column is itself a formula, and lands on this literal string
# for any row whose Order Date is blank (order not yet linked/resolved) - confirmed on both
# brands' live sheets, always paired with a blank Order Date. Ctx.cell() below swaps it for
# the row's own ticket month instead of surfacing the sentinel to callers.
ORDER_MONTH_SENTINEL = "12_Dec'99"

_SKU_TITLE_MAP = None


def sku_title_map():
    """SKU (casefold) -> canonical Product Name, from SKU master.xlsx's `Product variant
    SKU` -> `Product title` columns at the repo root. First-seen title wins per SKU (same
    convention as ci_key), so the handful of SKUs the master itself lists under more than
    one title resolve to whichever row came first rather than flapping between runs.
    mCaffeine-only: build_cross_filter_panel/the SKU drill-down engine only apply this to
    the "prod" column for brands whose Ctx passes one in (see Ctx.__init__)."""
    global _SKU_TITLE_MAP
    if _SKU_TITLE_MAP is not None:
        return _SKU_TITLE_MAP
    path = REPO_ROOT / "SKU master.xlsx"
    if not path.exists():
        _SKU_TITLE_MAP = {}
        return _SKU_TITLE_MAP
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    m = {}
    for title, sku in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
        if not sku or not title:
            continue
        m.setdefault(str(sku).strip().casefold(), str(title).strip())
    _SKU_TITLE_MAP = m
    return _SKU_TITLE_MAP


def h_enc(s):
    """Matches .NET Core's HttpUtility.HtmlEncode for the plain-ASCII case used here:
    escapes &, <, >, " and ' (as the numeric entity &#39;, not &apos;)."""
    s = "" if s is None else str(s)
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;").replace("'", "&#39;"))


def round1(n):
    return round(n, 1)


def j_enc(s):
    s = "" if s is None else str(s)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\r\n", "\\n").replace("\r", "\\n").replace("\n", "\\n").replace("\t", "\\t")
    return s


def pretty_month(raw):
    parts = raw.split("_", 1)
    if len(parts) < 2:
        return raw
    return parts[1].replace("'", " '")


_year_of_cache = {}
_year_re = re.compile(r"['\s](\d{2})$")


def year_of(mo):
    if mo in _year_of_cache:
        return _year_of_cache[mo]
    m = _year_re.search(mo)
    y = f"20{m.group(1)}" if m else ""
    _year_of_cache[mo] = y
    return y


_month_num_re = re.compile(r"^(\d+)_")


def parse_month_label(mo):
    """'7_Jul'26' -> (2026, 7), or None if unparseable. Sheet month labels always start
    with the month number and end with a 2-digit year (see year_of)."""
    m = _month_num_re.match(mo)
    yr = year_of(mo)
    if not m or not yr:
        return None
    return int(yr), int(m.group(1))


def nice_max(v):
    if v <= 0:
        return 10
    m = math.pow(10, math.floor(math.log10(v)))
    for s in (1, 2, 2.5, 5, 10):
        c = s * m
        if c >= v:
            return c
    return 10 * m


def fnum(v):
    """Approximates .NET's default double.ToString() for SVG coordinate/dimension values:
    no trailing '.0' for whole numbers (Python's str(float) always adds one), and rounds
    away floating-point noise so equivalent-but-differently-ordered arithmetic doesn't
    produce visually-identical-but-longer strings like 115.57000000000001 vs 115.57."""
    v = round(v, 6)
    if v == int(v):
        return str(int(v))
    s = f"{v:.6f}".rstrip("0").rstrip(".")
    return s


def index_map(seq):
    """{value: its index in seq} - the O(1) replacement for calling seq.index(v) inside a
    per-row loop, which is a linear scan of seq per row. The panel builders index ~100k
    ticket rows against lists of hundreds (delivery partners, query categories) to
    thousands (SKUs, batches) of entries.

    Worth roughly a second per brand, not more: list.index() is a tight C loop, so despite
    the O(n*k) shape it was never the bottleneck the sizes suggest. Kept because it's
    strictly cheaper and removes the misleading shape, not as a fix for anything.

    First occurrence wins (setdefault), so this matches list.index() exactly even if seq
    somehow holds a duplicate - e.g. dim2_order's appended "(other)" sentinel colliding
    with a real value literally named "(other)"."""
    d = {}
    for i, v in enumerate(seq):
        d.setdefault(v, i)
    return d


def n0(v):
    """Matches PowerShell's $n.ToString('N0') under this machine's culture, which formats
    with Indian digit grouping (lakh/crore: 1,67,751) rather than Western thousands
    grouping (167,751) - confirmed by diffing against the live PowerShell output."""
    iv = round(v)
    neg = iv < 0
    s = str(abs(iv))
    if len(s) <= 3:
        result = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        result = ",".join(parts) + "," + last3
    return ("-" if neg else "") + result


# $script:CatNormMap/$script:PartnerNormMap in the original are also plain @{} hashtables,
# so their .ContainsKey() lookups are case-insensitive too - matched here via a lowercased
# lookup table built once at import time.
_CAT_NORM_MAP_CI = {k.casefold(): v for k, v in CAT_NORM_MAP.items()}
_PARTNER_NORM_MAP_CI = {k.casefold(): v for k, v in PARTNER_NORM_MAP.items()}


class Ctx:
    """Per-brand shared state, built once by generate_report.py and passed to every
    panel-builder module - the direct replacement for PowerShell's shared script scope."""

    def __init__(self, brand):
        self.b = brand
        self.col = brand["col"]
        # Set from --quick by generate_report.py. Sources that can serve a run from their
        # on-disk cache instead of a live re-query check this. Defaults False so anything
        # constructing a bare Ctx (tests, ad-hoc scripts) gets full-refresh behaviour.
        self.quick = False
        # cell() is called tens of millions of times per brand, so the only two columns it
        # has to normalize are resolved to plain ints once here instead of being looked up
        # out of self.col on every call. -1 default: no caller ever passes a negative
        # index, so a brand missing either key simply never normalizes that column.
        self._cat_i = self.col.get("cat", -1)
        self._partner_i = self.col.get("partner", -1)
        # Product Name is free text a rep typed per ticket, so the same SKU accrues several
        # near-duplicate spellings over time (size suffix, rebrand, typo) - the "same SKU,
        # different product name" rows seen in the Product/Suggestion panels and the
        # SKU->Product drill-down. Swap in the SKU master's canonical title when the row's
        # SKU is in it, so every ticket for a SKU rolls up under one name; brands without a
        # master file (sku_title_map() empty) fall through to the raw cell value unchanged.
        self._prod_i = self.col.get("prod", -1)
        self._sku_i = self.col.get("sku", -1)
        self._sku_titles = sku_title_map() if brand.get("brand") == "mcaffeine" else {}
        # See ORDER_MONTH_SENTINEL above - falls back to the row's own ticket month.
        self._order_month_i = self.col.get("order_month", -1)
        self._month_i = self.col.get("month", -1)

    def cell(self, row, i):
        if row is None:
            return ""
        if isinstance(row, list):
            if i < len(row):
                v = row[i]
                if v is None:
                    return ""
            else:
                return ""
        elif i == 0:
            v = row
        else:
            return ""
        # Integer column check first, and casefold() computed once rather than twice: it
        # allocates a new string every time, and the overwhelming majority of calls are for
        # columns with no normalization map at all.
        if i == self._cat_i:
            if isinstance(v, str):
                return _CAT_NORM_MAP_CI.get(v.casefold(), v)
        elif i == self._partner_i:
            if isinstance(v, str):
                return _PARTNER_NORM_MAP_CI.get(v.casefold(), v)
        elif i == self._prod_i and self._sku_titles and isinstance(v, str) and v.strip():
            sku_val = row[self._sku_i] if isinstance(row, list) and 0 <= self._sku_i < len(row) else None
            if isinstance(sku_val, str):
                canon = self._sku_titles.get(sku_val.strip().casefold())
                if canon:
                    return canon
        elif i == self._order_month_i and v == ORDER_MONTH_SENTINEL:
            tm = row[self._month_i] if isinstance(row, list) and 0 <= self._month_i < len(row) else None
            if tm not in (None, ""):
                return tm
        return v

    def count_by(self, data, i):
        d = {}
        ci_cache = {}
        for r in data:
            v = self.cell(r, i)
            if not str(v).strip():
                v = "(blank)"
            v = ci_key(v, ci_cache)
            d[v] = d.get(v, 0) + 1
        return d

    @staticmethod
    def top_n(d, n):
        items = sorted(d.items(), key=lambda kv: kv[1], reverse=True)[:n]
        return [{"key": k, "value": v} for k, v in items]


def sort_keys_by_last_period(by_key, tot_by_key, period_order):
    """Row order for a pivot table: descending by each key's value in the LAST period
    (from period_order - a month/week list, chronological) that has ANY data across all
    keys, rather than by cumulative total across every period - so a row's position
    reflects what's currently happening, not stale historical volume that may not even
    be visible anymore (e.g. after empty leading months are hidden). Ties within that
    period fall back to descending-by-total, for a stable order among rows that are all
    zero (or equal) in the current period. Matches the ProdPkg drill-down's existing
    lmk()/LP convention (gen_panels.py), just generalized for any {key: {period: count}}
    breakdown.
    by_key: {key: {period: count}}. tot_by_key: {key: total} - defines the key set/tiebreak.
    period_order: chronological list of periods (e.g. ctx.months or a month's week_list).
    """
    last_period = None
    for p in reversed(period_order):
        if any(by_key.get(k, {}).get(p, 0) for k in tot_by_key):
            last_period = p
            break
    if last_period is None:
        return [k for k, _ in sorted(tot_by_key.items(), key=lambda kv: kv[1], reverse=True)]
    return [k for k, _ in sorted(
        tot_by_key.items(),
        key=lambda kv: (by_key.get(kv[0], {}).get(last_period, 0), kv[1]),
        reverse=True,
    )]
